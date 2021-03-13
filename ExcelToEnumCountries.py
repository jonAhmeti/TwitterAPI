import openpyxl
import os

xl = openpyxl.load_workbook('./worldcities.xlsx')

countriesTuple = set()


def refineName(name):
    words = str(name).split(' `')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('. ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('\'-')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('\' ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('. ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = str(name).split(',')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = str(name).split('/')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('-')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('\'')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('`')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split(' ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('._')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('(')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('__')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split(')')
    refinedName = words[0]

    return refinedName


for row in xl['Sheet1'].iter_rows():

    countryFullname = ''
    countryISO = ''
    for cell in row:
        if cell.col_idx == 3:
            countryFullname = cell.value
        if cell.col_idx == 4:
            countryISO = cell.value

    countriesTuple.add((countryISO, countryFullname))

print(countriesTuple)
for country in countriesTuple:
    with open(f'Locations\\Locations.py', 'w') as creator:
        creator.write(f'from enum import Enum\n')

citiesToPython = set()

for row in xl['Sheet1'].iter_rows():
    lat = ''
    lng = ''
    name = ''
    var = ''
    country = ''

    for cell in row:
        if cell.col_idx == 1:
            lat = str(cell.value)
        elif cell.col_idx == 2:
            lng = str(cell.value)
        elif cell.col_idx == 4:
            country = str(cell.value)
        elif cell.col_idx == 5:
            name = str(cell.value)
        elif cell.col_idx == 6:
            var = str(cell.value)
    celldata = f'    {refineName(var)} = (\'{lat},{lng}\', \"{name}\")$$${country}'
    citiesToPython.add(celldata)

with open('Locations\\Locations.py', 'a', encoding='UTF-8') as file:
    for element in countriesTuple:
        file.write(f'\n\nclass {element[0]}(Enum):\n    full_name = "{element[1]}"\n')
        for city in citiesToPython:
            if city.split('$$$')[1] == element[0]:
                file.write(f"{city.split('$$$')[0]}\n")

