import openpyxl
import os

xl = openpyxl.load_workbook('./worldcities.xlsx')

countriesTuple = set()


def refineName(name):
    words = str(name).split(' `')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('. ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('\'-')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('\' ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('. ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('-')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('\'')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split('`')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    words = refinedName.split(' ')
    refinedName = ''
    for word in words:
        if words.index(word) != len(words) - 1:
            refinedName += word + "_"
        else:
            refinedName += word

    return refinedName


for row in xl['Sheet1'].iter_rows():

    countryFullname = ''
    countryISO = ''
    for cell in row:
        if cell.col_idx == 3:
            countryFullname = cell.value
        if cell.col_idx == 4:
            countryISO = cell.value

    countriesTuple.add((countryISO, countryFullname))

print(countriesTuple)
for country in countriesTuple:
    with open(f'Locations\\{country[0]}.py', 'w') as creator:
        creator.write(f'class {country[0]}:\n')
        creator.write(f'    full_name = \"{country[1]}\"')

citiesToPython = set()

for row in xl['Sheet1'].iter_rows():
    lat = ''
    lng = ''
    name = ''
    var = ''
    country = ''

    for cell in row:
        if cell.col_idx == 1:
            lat = str(cell.value)
        elif cell.col_idx == 2:
            lng = str(cell.value)
        elif cell.col_idx == 4:
            country = str(cell.value)
        elif cell.col_idx == 5:
            name = str(cell.value)
        elif cell.col_idx == 6:
            var = str(cell.value)
    celldata = f'    {refineName(var)} = (\'{lat},{lng}\', \"{name}\")$$${country}'
    citiesToPython.add(celldata)

for file in os.listdir('Locations'):
    with open(f'Locations\\{file}', 'a', encoding='utf-8') as writer:
        for row in xl['Sheet1'].iter_rows():
            cell_iso = ''
            cell_var = ''
            for cell in row:
                if cell.col_idx == 4:
                    cell_iso = str(cell.value)
                elif cell.col_idx == 6:
                    cell_var = str(cell.value)

            if cell_iso == file.split('.')[0]:
                for el in citiesToPython:
                    twoPart = el.split('$$$')
                    if el.split(' =')[0].lstrip() == refineName(cell_var) and twoPart[1] == cell_iso:
                        writer.write(f'\n{twoPart[0]}')
